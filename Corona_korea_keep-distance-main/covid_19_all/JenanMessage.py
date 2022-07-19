# 재난문자 API 받아오기
# 이제 재난문자를 서버내의 JenanMessage.xlsx에 저장합니다.

import requests
import xmltodict
import json
import openpyxl

# 코로나 관련 재난문자 전체 지역 저장
def jenan_all():
    # url의 numOfRows 파라미터 조절시 한번에 보여지는 양 설정, 많을 시 로딩시간 오래걸림
    print("재난문자를 로딩중입니다. (최대 1분 소요)")
    try:
        url = 'http://apis.data.go.kr/1741000/DisasterMsg3/getDisasterMsg1List?serviceKey=tL8DvlXmKWq0V7ralHks5bdaNOVJ4Y1yMkYncaEWfjTO%2F3bobA%2FuSCSDuVBxesTC%2F3lbC8JcFJZJJe9j9GoPgQ%3D%3D&pageNo=1&numOfRows=1000&type=xml'

        res = requests.get(url)
        if res.status_code == 200:
            # Ordered dictionary type
            result = xmltodict.parse(res.text)
            dd = json.loads(json.dumps(result))
            print('Road OK')
        else:
            print('res.status_code is NOT ok')

        wb = openpyxl.load_workbook('JenanMessage.xlsx')

        # 이전 내용 삭제
        while wb.active.max_row > 1:
            wb.active.delete_rows(2)

        i = 0
        for rgn in dd['DisasterMsg']['row']:
            if dd['DisasterMsg']['row'][i]['msg'].find("검사") != -1:
                # 엑셀 저장
                wb.active.append([dd['DisasterMsg']['row'][i]['create_date'],dd['DisasterMsg']['row'][i]['location_name'],dd['DisasterMsg']['row'][i]['msg']])
                wb.save('JenanMessage.xlsx')
            i += 1
        print("Save OK")
    except:
        print("Save Fail")


# 해당지역 재난문자 불러오기
def jenan_area(area):
    wb = openpyxl.load_workbook('JenanMessage.xlsx')

    cnt = 0
    print(area, "지역의 최신 코로나 관련 재난문자입니다.")
    for i in wb.active.rows:
        if area[0:2] in str(i[1].value) or area[0:2] in str(i[2].value):
            cnt += 1
            print("발령시간 : ", i[0].value)
            print("지역 : ", i[1].value)
            print("내용 : ", i[2].value)
    if cnt == 0:
        print(area,"지역에 대한 최근 코로나 관련 재난문자가 없습니다. \n")


if __name__ == "__main__":
    print("로딩중... ")
    # jenan_all()
    # jenan_area("대구")

jenan_all()
#jenan_area("춘천시")