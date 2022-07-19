# 소분류 지역별 코로나 확진자
import operator

import requests
from bs4 import BeautifulSoup

# 매일 GPS 기록 저장
import pandas as pd
import datetime

import re

import openpyxl

today = datetime.datetime.now()
yesterday = today - datetime.timedelta(1)
today = today.strftime("%Y%m%d")
yesterday = yesterday.strftime("%Y%m%d")

wb = openpyxl.load_workbook('CovidCount.xlsx')


def save():
    try:
        url = requests.get(
            'http://ncov.mohw.go.kr/bdBoardList_Real.do?brdId=1&brdGubun=13&ncvContSeq=&contSeq=&board_id=&gubun=')
        CovidCountURL = BeautifulSoup(url.content, "html.parser")
        today = datetime.datetime.now()
        yesterday = today - datetime.timedelta(1)
        today = today.strftime("%Y%m%d")
        yesterday = yesterday.strftime("%Y%m%d")

        wb = openpyxl.load_workbook('CovidCount.xlsx')
        # 이부분에 이전 Today자료를 YesterDay시트로 옮겨줘야함
        print("날짜가 바뀌어 데이터를 저장합니다.")
        # today -> yesterday , 오늘 새로운 데이터 추가

        wb.create_sheet(index=0, title=today)
        for i in range(0, 18, 1):
            regionData = CovidCountURL.select("#zone_popup" + str(i) + " > div > table > tbody > tr")

            for rgn in regionData:
                # 콤마 제거
                rgn = rgn.text.replace(',', '')
                data = {
                    'date': [today],
                    'region': [rgn],
                }
                # 엑셀 저장
                td_sheet = wb[today]
                td_sheet.append([rgn])
                wb.save('CovidCount.xlsx')
        print("Save OK")
    except:
        print("Save Error")


def CovidCountSave():
    today = datetime.datetime.now()
    yesterday = today - datetime.timedelta(1)
    today = today.strftime("%Y%m%d")
    yesterday = yesterday.strftime("%Y%m%d")

    try:
        # 데이터 파일의 첫번째 인덱스 시트의 값과 오늘 날짜 비교, 다르면 저장
        if int(wb.sheetnames[0]) != int(today):
            save()
            state = 1
        else:
            # 이미 저장된 자료가 있음
            print("이전과 같음")
            state = 0

    except:
        save()
        state = -1

    # state -1 : 오류(파일없음 등) / 0 : 데이터의 변화 없음 저장안함 / 1 : 새로운 데이터를 저장함
    print(state, datetime.datetime.now())
    return state


def find(area):
    try:
        td_sheet = wb[today]  # 오늘 시트
        ys_sheet = wb[yesterday]  # 전날 시트
    except:
        # 오늘 또는 전날 기록이 없어 부정확한 자료입니다.
        td_sheet = wb[wb.sheetnames[0]]
        ys_sheet = wb[wb.sheetnames[1]]

    for i in ys_sheet.rows:
        if area in i[0].value:
            ys_cnt = int(re.findall("\d+", i[0].value)[0])
            break

    for i in td_sheet.rows:
        if area in i[0].value:
            td_cnt = int(re.findall("\d+", i[0].value)[0])
            break

    # 오늘 누적 확진자와 어제 누적 확진자의 차이
    cnt_today = td_cnt - ys_cnt
    return cnt_today


# 위험지역 확인
def findTopDanger(count=10):
    wb = openpyxl.load_workbook('CovidCount.xlsx')
    try:
        td_sheet = wb[today]  # 오늘 시트
        ys_sheet = wb[yesterday]  # 전날 시트
    except:
        # 오늘 또는 전날 기록이 없어 부정확한 자료입니다.
        td_sheet = wb[wb.sheetnames[0]]
        ys_sheet = wb[wb.sheetnames[1]]

    td_cnt = []
    ys_cnt = []
    todayIncrease = []
    population = []
    incidenceRate = []  # 발생률
    areaLen = 249

    # 지역별 발생 수
    for i in td_sheet.rows:
        td_cnt.append(int(re.findall("\d+", i[0].value)[0]))
    for i in ys_sheet.rows:
        ys_cnt.append(int(re.findall("\d+", i[0].value)[0]))

    # 증가 수치
    for i in range(areaLen):
        todayIncrease.append((td_cnt[i] - ys_cnt[i]))

    population_sheet = wb[wb.sheetnames[-1]]

    for i in population_sheet.rows:
        population.append(i[2].value)

    for i in range(areaLen):  # 10000명당 발생률
        incidenceRate.append(todayIncrease[i] / population[i] * 10000)

    tempIncidenceRate = incidenceRate
    tempIncidenceRate = sorted(tempIncidenceRate)
    tempIncidenceRate.reverse()

    tempIndex = []
    for i in incidenceRate:
        tempIndex.append(tempIncidenceRate.index(i) + 1)

    # 가장 많은 발생 count 갯수만큼 반환
    # 도 / 시,구 / 만명당 발생률

    returnMsg = []
    for i in range(0, count):
        returnMsg.append("""
        <article class="post">
            <header>
                <div class="meta">
                     <h2>{:}</h2>
                     <h2>{:}</h2>
                     <h3>{:.2f}</h3>
                </div>
            </header>
            </article>
        """.format(population_sheet[tempIndex.index(i + 1)][0].value,
                   population_sheet[tempIndex.index(i + 1)][1].value,
                   incidenceRate[tempIndex.index(i + 1)]))

    return returnMsg

# 비위험지역 확인
def findLowDanger(count=10):
    wb = openpyxl.load_workbook('CovidCount.xlsx')
    try:
        td_sheet = wb[today]  # 오늘 시트
    except:
        td_sheet = wb[wb.sheetnames[0]]
    try:
        ys_sheet = wb[yesterday]  # 전날 시트
    except:
        # 전날 기록이 없어 부정확한 자료입니다.
        ys_sheet = wb[wb.sheetnames[1]]

    td_cnt = []
    ys_cnt = []
    todayIncrease = []
    population = []
    incidenceRate = []  # 발생률
    areaLen = 249

    # 지역별 발생 수
    for i in td_sheet.rows:
        td_cnt.append(int(re.findall("\d+", i[0].value)[0]))
    for i in ys_sheet.rows:
        ys_cnt.append(int(re.findall("\d+", i[0].value)[0]))

    # 증가 수치
    for i in range(areaLen):
        todayIncrease.append((td_cnt[i] - ys_cnt[i]))

    population_sheet = wb[wb.sheetnames[-1]]

    for i in population_sheet.rows:
        population.append(i[2].value)

    zeroIncrease = []
    for i in range(areaLen):  # 10000명당 발생률
        if todayIncrease[i] == 0: # 0명 발생지역
            zeroIncrease.append(population[i])

    TempzeroIncrease = zeroIncrease
    TempzeroIncrease = sorted(TempzeroIncrease)
    TempzeroIncrease.reverse()


    # 가장 많은 발생 count 갯수만큼 반환
    # 도 / 시,구 / 만명당 발생률

    returnMsg = []
    TempzeroIncrease[0]


    for T in TempzeroIncrease:
        for i in wb.active.rows:
            if str(T) in str(i[2].value):
                returnMsg.append("""
                <article class="post">
                    <header>
                        <div class="meta">
                             <h2>{:}</h2>
                             <h2>{:}</h2>
                             <h3>{:}</h3>
                        </div>
                    </header>
                    </article>
                """.format(i[0].value, i[1].value,
                           i[2].value))

    return returnMsg

def __init__():
    CovidCountSave()


# CovidCountSave()
# print(find("원주시"))
# print(find("춘천시"))
''' 
    서버에선 다음을 실행해야함 
    CovidCountSave()
'''
'''
    사용자는 다음을 실행해야함
    find(area)
    #print(find("원주시"))
    #print(find("춘천시"))
'''
