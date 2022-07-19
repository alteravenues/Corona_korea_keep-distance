from django import template#html에서 로드하기 위한 부분
register = template.Library()#html에서 로드하기 위한 부분
from openpyxl import load_workbook

import os
@register.simple_tag#html에서 로드하기 위한 부분
def load_excel():
    BASE_DIR = os.getcwd()
    load_wb = load_workbook(str(BASE_DIR)+"\\corona\\templatetags\\행정구역.xlsx",data_only=True)
    load_ws = load_wb['KIKcd_B']

    cell_all = []
    cell_count = 1

    while 1:
        a = load_ws.cell(cell_count,1).value
        b = load_ws.cell(cell_count,2).value
        if(a==None):
            break
        cell_all.append([str(a),str(b)])
        cell_count+=1
    return cell_all
@register.simple_tag#html에서 로드하기 위한 부분

def kkk(location):
    return location